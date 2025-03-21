import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "issues.xlsx"
SHEET_NAME = "Issues"

def update_excel(issue_data):
    try:
        # Try loading the existing Excel file
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            sheet = wb[SHEET_NAME]
        else:
            raise FileNotFoundError

    except (FileNotFoundError, OSError):
        # Create a new workbook if the file doesn't exist or is invalid
        wb = Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME
        # Add headers to the new workbook
        sheet.append(["Issue ID", "Title", "State", "Created At", "Labels"])

    # Append new issue data
    sheet.append([
        issue_data["id"],
        issue_data["title"],
        issue_data["state"],
        issue_data["created_at"],
        ", ".join(issue_data["labels"])
    ])

    # Save the Excel file
    wb.save(EXCEL_FILE)

if __name__ == "__main__":
    issue_data = {
        "id": os.environ["ISSUE_ID"],
        "title": os.environ["ISSUE_TITLE"],
        "state": os.environ["ISSUE_STATE"],
        "created_at": os.environ["ISSUE_CREATED_AT"],
        "labels": os.environ.get("ISSUE_LABELS", "").split(",") if os.environ.get("ISSUE_LABELS") else []
    }
    update_excel(issue_data)
